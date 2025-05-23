from django.shortcuts import render, get_object_or_404, redirect
from .models import Car
from django.urls import reverse_lazy
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from .models import Cliente
from django.db.models import Q
from django.forms import ModelForm
from django import forms
from django.http import HttpResponse
from openpyxl import Workbook
from django.template.loader import get_template
from xhtml2pdf import pisa
from io import BytesIO

def export_cars_excel(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="inventario_carros.xlsx"'
    
    # Aplicar los mismos filtros de búsqueda
    queryset = Car.objects.all()
    search_query = request.GET.get('search', '')
    disponible = request.GET.get('disponible', '')
    precio_min = request.GET.get('precio_min', '')
    precio_max = request.GET.get('precio_max', '')

    if search_query:
        queryset = queryset.filter(
            Q(marca__icontains=search_query) |
            Q(modelo__icontains=search_query) |
            Q(año__icontains=search_query) |
            Q(cliente__nombre__icontains=search_query) |
            Q(cliente__apellido__icontains=search_query)
        )

    if disponible:
        queryset = queryset.filter(disponible=disponible == 'true')

    if precio_min:
        queryset = queryset.filter(precio__gte=precio_min)
    
    if precio_max:
        queryset = queryset.filter(precio__lte=precio_max)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario de Carros"
    
    # Headers
    columns = ['Cliente', 'Marca', 'Modelo', 'Año', 'Precio', 'Color', 'Kilometraje', 'Estado', 'Fecha Ingreso']
    for col_num, column_title in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = column_title
    
    # Data
    for row_num, car in enumerate(queryset, 2):
        ws.cell(row=row_num, column=1, value=str(car.cliente) if car.cliente else "Sin asignar")
        ws.cell(row=row_num, column=2, value=car.marca)
        ws.cell(row=row_num, column=3, value=car.modelo)
        ws.cell(row=row_num, column=4, value=car.año)
        ws.cell(row=row_num, column=5, value=float(car.precio))
        ws.cell(row=row_num, column=6, value=car.color)
        ws.cell(row=row_num, column=7, value=car.kilometraje)
        ws.cell(row=row_num, column=8, value="Disponible" if car.disponible else "No Disponible")
        ws.cell(row=row_num, column=9, value=car.fecha_ingreso.strftime("%d/%m/%Y"))
    
    wb.save(response)
    return response

def export_cars_pdf(request):
    # Aplicar los mismos filtros de búsqueda
    queryset = Car.objects.all()
    search_query = request.GET.get('search', '')
    disponible = request.GET.get('disponible', '')
    precio_min = request.GET.get('precio_min', '')
    precio_max = request.GET.get('precio_max', '')

    if search_query:
        queryset = queryset.filter(
            Q(marca__icontains=search_query) |
            Q(modelo__icontains=search_query) |
            Q(año__icontains=search_query) |
            Q(cliente__nombre__icontains=search_query) |
            Q(cliente__apellido__icontains=search_query)
        )

    if disponible:
        queryset = queryset.filter(disponible=disponible == 'true')

    if precio_min:
        queryset = queryset.filter(precio__gte=precio_min)
    
    if precio_max:
        queryset = queryset.filter(precio__lte=precio_max)

    template = get_template('inventory/car_pdf.html')
    context = {
        'cars': queryset,
        'request': request,
        'search_query': search_query
    }
    html = template.render(context)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="inventario_carros.pdf"'
    
    # Create PDF
    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('Error al generar el PDF', status=500)
    return response





class CarForm(ModelForm):
    class Meta:
        model = Car
        fields = ['cliente', 'marca', 'modelo', 'año', 'precio', 'color', 'kilometraje', 'disponible']
        widgets = {
            'cliente': forms.Select(attrs={'class': 'form-select'}),
            'marca': forms.TextInput(attrs={'class': 'form-control'}),
            'modelo': forms.TextInput(attrs={'class': 'form-control'}),
            'año': forms.NumberInput(attrs={'class': 'form-control'}),
            'precio': forms.NumberInput(attrs={'class': 'form-control'}),
            'color': forms.TextInput(attrs={'class': 'form-control'}),
            'kilometraje': forms.NumberInput(attrs={'class': 'form-control'}),
            'disponible': forms.CheckboxInput(attrs={'class': 'form-check-input'})
        }



class CarListView(ListView):
    model = Car
    template_name = 'inventory/car_list.html'
    context_object_name = 'cars'
    
    def get_queryset(self):
        queryset = super().get_queryset()
        search_query = self.request.GET.get('search', '')
        disponible = self.request.GET.get('disponible', '')
        precio_min = self.request.GET.get('precio_min', '')
        precio_max = self.request.GET.get('precio_max', '')

        if search_query:
            queryset = queryset.filter(
                Q(marca__icontains=search_query) |
                Q(modelo__icontains=search_query) |
                Q(año__icontains=search_query) |
                Q(cliente__nombre__icontains=search_query) |
                Q(cliente__apellido__icontains=search_query)
            )

        if disponible:
            queryset = queryset.filter(disponible=disponible == 'true')

        if precio_min:
            queryset = queryset.filter(precio__gte=precio_min)
        
        if precio_max:
            queryset = queryset.filter(precio__lte=precio_max)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search_query'] = self.request.GET.get('search', '')
        context['disponible'] = self.request.GET.get('disponible', '')
        context['precio_min'] = self.request.GET.get('precio_min', '')
        context['precio_max'] = self.request.GET.get('precio_max', '')
        return context

class CarCreateView(CreateView):
    model = Car
    template_name = 'inventory/car_form.html'
    form_class = CarForm
    success_url = reverse_lazy('car_list')

class CarUpdateView(UpdateView):
    model = Car
    template_name = 'inventory/car_form.html'
    form_class = CarForm
    success_url = reverse_lazy('car_list')

class CarDeleteView(DeleteView):
    model = Car
    template_name = 'inventory/car_confirm_delete.html'
    success_url = reverse_lazy('car_list')
    

class ClienteListView(ListView):
    model = Cliente
    template_name = 'inventory/cliente_list.html'
    context_object_name = 'clientes'
    
    def get_queryset(self):
        queryset = super().get_queryset()
        search_query = self.request.GET.get('search', '')
        activo = self.request.GET.get('activo', '')

        if search_query:
            queryset = queryset.filter(
                Q(nombre__icontains=search_query) |
                Q(apellido__icontains=search_query) |
                Q(email__icontains=search_query) |
                Q(telefono__icontains=search_query)
            )

        if activo:
            queryset = queryset.filter(activo=activo == 'true')

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search_query'] = self.request.GET.get('search', '')
        context['activo'] = self.request.GET.get('activo', '')
        return context
    

class ClienteCreateView(CreateView):
    model = Cliente
    template_name = 'inventory/cliente_form.html'
    fields = ['nombre', 'apellido', 'email', 'telefono', 'direccion', 'activo']
    success_url = reverse_lazy('cliente_list')

class ClienteUpdateView(UpdateView):
    model = Cliente
    template_name = 'inventory/cliente_form.html'
    fields = ['nombre', 'apellido', 'email', 'telefono', 'direccion', 'activo']
    success_url = reverse_lazy('cliente_list')

class ClienteDeleteView(DeleteView):
    model = Cliente
    template_name = 'inventory/cliente_confirm_delete.html'
    success_url = reverse_lazy('cliente_list')